import openpyxl
from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Font, PatternFill

from .models import (
    Club,
    ClubLineaInvestigacion,
    Dependencia,
    Estado,
    Institucion,
    LineaInvestigacion,
    MembresiaClu,
    Municipio,
    Parroquia,
    Participante,
)


@admin.register(Estado)
class EstadoAdmin(admin.ModelAdmin):
    list_display = ("id", "nombre", "codigo")
    search_fields = ("nombre", "codigo")
    ordering = ("nombre",)


@admin.register(Dependencia)
class DependenciaAdmin(admin.ModelAdmin):
    list_display = ("id", "nombre", "activa")
    search_fields = ("nombre",)
    list_filter = ("activa",)


@admin.register(Municipio)
class MunicipioAdmin(admin.ModelAdmin):
    list_display = ("nombre", "estado")
    list_filter = ("estado",)  # Filtro lateral por estado
    search_fields = ("nombre",)
    ordering = ("estado", "nombre")
    # Permite que Parroquia busque municipios de forma asíncrona
    search_fields = ["nombre"]


@admin.register(Parroquia)
class ParroquiaAdmin(admin.ModelAdmin):
    list_display = ("nombre", "get_municipio", "get_estado")
    search_fields = ("nombre",)

    # Mejora de rendimiento: Filtra por el estado del municipio
    list_filter = ("municipio__estado",)

    # Nivel Pro: Autocompletado para no colapsar el navegador con miles de opciones
    autocomplete_fields = ["municipio"]

    @admin.display(description="Municipio", ordering="municipio__nombre")
    def get_municipio(self, obj):
        return obj.municipio.nombre

    @admin.display(description="Estado", ordering="municipio__estado__nombre")
    def get_estado(self, obj):
        return obj.municipio.estado.nombre


@admin.register(Participante)
class ParticipanteAdmin(admin.ModelAdmin):
    list_display = [
        "mostrar_cedula",
        "nombres",
        "apellidos",
        "email",
        "estado",
        "fecha_registro",
    ]
    list_filter = ["estado", "sexo", "grado_escolar", "condicion_tea", "nacionalidad"]
    search_fields = ["cedula", "cedula_escolar", "nombres", "apellidos", "email"]
    readonly_fields = ["fecha_registro"]
    ordering = ["-fecha_registro"]

    @admin.display(description="Cédula", ordering="cedula")
    def mostrar_cedula(self, obj):
        """Muestra cédula personal con prioridad, o cédula escolar si no tiene personal."""
        if obj.cedula:
            return f"{obj.nacionalidad}-{obj.cedula}"
        elif obj.cedula_escolar:
            return f"CE-{obj.cedula_escolar}"
        return "Sin cédula"

    fieldsets = (
        (
            "Identificación",
            {
                "fields": (
                    "nacionalidad",
                    "cedula",
                    "cedula_escolar",
                    "condicion_tea",
                )
            },
        ),
        (
            "Datos Personales",
            {
                "fields": (
                    "nombres",
                    "apellidos",
                    "fecha_nacimiento",
                    "sexo",
                    "email",
                    "codigo_area",
                    "numero_telefono",
                    "direccion",
                )
            },
        ),
        (
            "Ubicación",
            {"fields": ("estado", "municipio", "parroquia")},
        ),
        (
            "Educación",
            {
                "fields": (
                    "grado_escolar",
                    "titulo_universitario",
                    "campo1",
                )
            },
        ),
        (
            "Representante (para menores de edad)",
            {
                "fields": (
                    "nombre_representante",
                    "nacionalidad_representante",
                    "cedula_representante",
                    "codigo_area_representante",
                    "numero_telefono_representante",
                    "email_representante",
                ),
                "classes": ("collapse",),
            },
        ),
        (
            "Metadata",
            {
                "fields": ("fecha_registro", "user"),
                "classes": ("collapse",),
            },
        ),
    )


# 1. Definimos la acción
@admin.action(description="Aprobar y generar código RNR")
def aprobar_registros(modeladmin, request, queryset):
    from .services.admission_service import AdmissionService

    count = 0
    for institucion in queryset.filter(estatus="pendiente"):
        if AdmissionService.approve_institution(institucion, request.user):
            count += 1
    modeladmin.message_user(
        request, f"Se han aprobado {count} instituciones correctamente."
    )


@admin.register(Institucion)
class InstitucionAdmin(admin.ModelAdmin):
    list_display = (
        "codigo",
        "nombre",
        "tipo_institucion",
        "naturaleza",
        "rif",
        "mostrar_cedula_particular",
        "email",
        "estado",
        "activa",
        "federado",
        "eliminado",
    )
    readonly_fields = ("codigo", "fecha_registro")
    exclude = ("tipo_federado",)
    list_filter = (
        "estatus",
        "activa",
        "federado",
        "eliminado",
        "estado",
        "tipo_institucion",
        "naturaleza",
    )
    search_fields = ("nombre", "codigo", "email", "rif", "codigo_mppe")
    actions = ["aprobar_instituciones", "exportar_excel"]

    fieldsets = (
        ("Identificación del Sistema", {"fields": ("codigo", "fecha_registro")}),
        (
            "Datos de Identificación Institucional",
            {
                "fields": (
                    "tipo_institucion",
                    "naturaleza",
                    "subcategoria",
                    "dependencia",
                    "dependencia_rel",
                )
            },
        ),
        (
            "Información General",
            {
                "fields": (
                    "nombre",
                    "rif",
                    "codigo_mppe",
                    "codigo_infocentro",
                    "federado",
                    "email",
                    "telefono",
                    "particular_nacionalidad",
                    "particular_cedula",
                )
            },
        ),
        (
            "Ubicación Geográfica",
            {"fields": ("estado", "municipio", "parroquia", "direccion")},
        ),
        ("Estado de la Cuenta", {"fields": ("estatus", "activa", "eliminado")}),
    )

    def aprobar_instituciones(self, request, queryset):
        from .services.admission_service import AdmissionService

        count = 0
        for inst in queryset.filter(estatus="pendiente"):
            if AdmissionService.approve_institution(inst, request.user):
                count += 1
        self.message_user(
            request,
            f"{count} instituciones han sido aprobadas exitosamente.",
        )

    aprobar_instituciones.short_description = "✅ Aprobar y generar códigos RNR"

    @admin.display(description="Céd.", ordering="particular_cedula")
    def mostrar_cedula_particular(self, obj):
        if obj.particular_nacionalidad and obj.particular_cedula:
            return f"{obj.particular_nacionalidad}-{obj.particular_cedula}"
        return ""

    def exportar_excel(self, request, queryset):
        # Asegurar que los objetos relacionados estén cargados para evitar N+1 queries
        queryset = queryset.select_related(
            "estado", "municipio", "parroquia", "dependencia_rel"
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Instituciones"

        # Encabezados
        headers = [
            "Código",
            "Nombre",
            "Tipo",
            "Naturaleza",
            "Subcategoría",
            "Dependencia",
            "RIF",
            "Código MPPE",
            "Email",
            "Teléfono",
            "Estado",
            "Municipio",
            "Parroquia",
            "Dirección",
            "Activa",
            "Federado",
            "Estatus",
            "Fecha Registro",
        ]
        ws.append(headers)

        # Estilo encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )

        # Datos
        for inst in queryset:
            # Obtener nombre de dependencia
            if inst.dependencia:
                dependencia_nombre = inst.dependencia
            elif inst.dependencia_rel:
                dependencia_nombre = inst.dependencia_rel.nombre
            else:
                dependencia_nombre = ""

            # Obtener teléfono completo
            telefono_completo = ""
            if inst.telefono_codigo and inst.telefono_numero:
                telefono_completo = f"{inst.telefono_codigo}-{inst.telefono_numero}"
            elif inst.telefono:
                telefono_completo = inst.telefono

            ws.append(
                [
                    inst.codigo,
                    inst.nombre,
                    inst.get_tipo_institucion_display()
                    if inst.tipo_institucion
                    else "",
                    inst.get_naturaleza_display() if inst.naturaleza else "",
                    inst.subcategoria or "",
                    dependencia_nombre,
                    inst.rif or "",
                    inst.codigo_mppe or "",
                    inst.email,
                    telefono_completo,
                    str(inst.estado) if inst.estado else "",
                    str(inst.municipio) if inst.municipio else "",
                    str(inst.parroquia) if inst.parroquia else "",
                    inst.direccion or "",
                    "Sí" if inst.activa else "No",
                    "Sí" if inst.federado else "No",
                    inst.get_estatus_display() if inst.estatus else "",
                    inst.fecha_registro.strftime("%d/%m/%Y %H:%M")
                    if inst.fecha_registro
                    else "",
                ]
            )

        # Ajustar ancho columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Respuesta HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="instituciones_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        )
        wb.save(response)
        return response

    exportar_excel.short_description = "⬇️ Exportar a Excel"


# =======================
# ADMIN DE LÍNEAS DE INVESTIGACIÓN
# =======================


@admin.register(LineaInvestigacion)
class LineaInvestigacionAdmin(admin.ModelAdmin):
    """Admin para gestionar Líneas de Investigación (Catálogo Dinámico)."""

    list_display = ["codigo", "nombre", "activa", "orden", "fecha_creacion"]
    list_filter = ["activa"]
    search_fields = ["codigo", "nombre", "descripcion"]
    list_editable = ["activa", "orden"]
    ordering = ["orden", "nombre"]

    fieldsets = (
        ("Información Básica", {"fields": ("codigo", "nombre", "descripcion")}),
        ("Configuración", {"fields": ("activa", "orden")}),
        (
            "Fechas",
            {
                "fields": ("fecha_creacion", "fecha_actualizacion"),
                "classes": ("collapse",),
            },
        ),
    )
    readonly_fields = ["fecha_creacion", "fecha_actualizacion"]


class ClubLineaInvestigacionInline(admin.TabularInline):
    """Inline para gestionar líneas de investigación de un club."""

    model = ClubLineaInvestigacion
    extra = 1
    max_num = 3
    min_num = 1
    fields = ["linea", "tipo_linea", "orden"]
    autocomplete_fields = ["linea"]


# =======================
# ADMIN DE CLUBES
# =======================


@admin.register(Club)
class ClubAdmin(admin.ModelAdmin):
    """Admin para gestionar Clubes."""

    list_display = [
        "nombre",
        "siglas",
        "institucion_creadora",
        "status",
        "estado_vinculacion",
        "cupo_maximo",
        "cupos_disponibles",
        "activo",
        "fecha_creacion",
    ]
    list_filter = [
        "status",
        "estado_vinculacion",
        "activo",
        "institucion_creadora__estado",
    ]
    search_fields = ["nombre", "siglas", "descripcion"]
    readonly_fields = ["fecha_creacion", "fecha_aprobacion"]
    inlines = [ClubLineaInvestigacionInline]

    fieldsets = (
        (
            "Información Básica",
            {"fields": ("nombre", "siglas", "logo", "descripcion", "ubicacion")},
        ),
        (
            "Institución Creadora",
            {"fields": ("institucion_creadora", "coordinador", "documento_legal")},
        ),
        (
            "Configuración de Vinculación",
            {"fields": ("estado_vinculacion", "cupo_maximo", "requisitos")},
        ),
        (
            "Estado y Fechas",
            {"fields": ("status", "activo", "fecha_creacion", "fecha_aprobacion")},
        ),
        # NOTA: Campos linea_1, linea_2, linea_3 eliminados - usar ClubLineaInvestigacionInline
    )

    actions = ["aprobar_clubes", "rechazar_clubes", "cerrar_clubes", "abrir_clubes"]

    def aprobar_clubes(self, request, queryset):
        from django.utils import timezone

        count = 0
        for club in queryset.filter(status__in=["pendiente", "en_revision"]):
            club.status = "aprobado"
            club.fecha_aprobacion = timezone.now()
            club.save(update_fields=["status", "fecha_aprobacion"])
            count += 1
        self.message_user(request, f"{count} clubes han sido aprobados.")

    aprobar_clubes.short_description = "✅ Aprobar clubes seleccionados"

    def rechazar_clubes(self, request, queryset):
        count = 0
        for club in queryset.filter(status__in=["pendiente", "en_revision"]):
            club.status = "rechazado"
            club.save(update_fields=["status"])
            count += 1
        self.message_user(request, f"{count} clubes han sido rechazados.")

    rechazar_clubes.short_description = "❌ Rechazar clubes seleccionados"

    def cerrar_clubes(self, request, queryset):
        count = 0
        for club in queryset.filter(estado_vinculacion="abierto"):
            club.estado_vinculacion = "cerrado"
            club.save(update_fields=["estado_vinculacion"])
            count += 1
        self.message_user(request, f"{count} clubes han sido cerrados.")

    cerrar_clubes.short_description = "🔒 Cerrar clubes (no acepta más postulaciones)"

    def abrir_clubes(self, request, queryset):
        count = 0
        for club in queryset.filter(estado_vinculacion="cerrado"):
            club.estado_vinculacion = "abierto"
            club.save(update_fields=["estado_vinculacion"])
            count += 1
        self.message_user(request, f"{count} clubes han sido abiertos.")

    abrir_clubes.short_description = "🔓 Abrir clubes (acepta postulaciones)"


@admin.register(MembresiaClu)
class MembresiaCluAdmin(admin.ModelAdmin):
    """Admin para gestionar Membresías de Clubes."""

    list_display = [
        "club",
        "institucion",
        "tipo_linea",
        "estado",
        "representante_legal",
        "fecha_solicitud",
        "fecha_respuesta",
    ]
    list_filter = ["estado", "tipo_linea", "club"]
    search_fields = ["club__nombre", "institucion__nombre", "representante_legal"]
    readonly_fields = ["fecha_solicitud"]

    fieldsets = (
        ("Club e Institución", {"fields": ("club", "institucion")}),
        (
            "Información de la Solicitud",
            {
                "fields": (
                    "carta_intencion",
                    "propuesta_tecnica",
                    "representante_legal",
                    "tipo_linea",
                )
            },
        ),
        (
            "Estado",
            {
                "fields": (
                    "estado",
                    "observaciones",
                    "fecha_solicitud",
                    "fecha_respuesta",
                )
            },
        ),
    )

    actions = ["aprobar_membresias", "rechazar_membresias"]

    def aprobar_membresias(self, request, queryset):
        from django.utils import timezone

        count = 0
        for membresia in queryset.filter(
            estado__in=["pendiente_filtro", "visto_bueno_fundadora"]
        ):
            membresia.estado = "miembro_activo"
            membresia.fecha_respuesta = timezone.now()
            membresia.save(update_fields=["estado", "fecha_respuesta"])
            # Guardar el club para actualizar cupos
            membresia.club.save()
            count += 1
        self.message_user(request, f"{count} membresías han sido aprobadas.")

    aprobar_membresias.short_description = "✅ Aprobar membresías seleccionadas"

    def rechazar_membresias(self, request, queryset):
        from django.utils import timezone

        count = 0
        for membresia in queryset.filter(
            estado__in=["pendiente_filtro", "visto_bueno_fundadora"]
        ):
            membresia.estado = "rechazada"
            membresia.fecha_respuesta = timezone.now()
            membresia.save(update_fields=["estado", "fecha_respuesta"])
            count += 1
        self.message_user(request, f"{count} membresías han sido rechazadas.")

    rechazar_membresias.short_description = "❌ Rechazar membresías seleccionadas"
